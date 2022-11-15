import PyPDF2
import camelot
import ctypes
import numpy as np
from ctypes.util import find_library
import pandas as pd
from tabulate import tabulate
import pdf2image
import xlsxwriter
import pytesseract
from pytesseract import Output, TesseractError
import cv2
import PIL
import os
import shutil
import re
from PIL import Image
from numpy import asarray
from matplotlib import pyplot as plt
import pdfplumber
import datetime
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'
poppler_path = r"C:\Program Files (x86)\poppler-22.04.0\Library\bin"

def pdf2txt(pdf_path):
    file = open(pdf_path, 'rb')
    fileReader = PyPDF2.PdfFileReader(file)
    num_pages = fileReader.numPages
    text_list = []
    for i in range(0, num_pages):
        text = str(fileReader.getPage(i).extract_text()).replace("\n", "").split(" ")
        while ("" in text):
            text.remove("")
        text_list.append(text)
    return text_list
def pdf2tbl(pdf_path, page):
    a = camelot.read_pdf(pdf_path, pages=str(page))
    # print(a[2].parsing_report)
    # print(a[2].df.to_excel("test.xlsx"))
    return a
def pdfimg2text(pdf_path_ocr, page):
    images = pdf2image.convert_from_path(pdf_path=pdf_path_ocr, poppler_path=poppler_path, first_page=page, last_page=page)
    img = asarray(images[0])
    img = cv2.resize(img, None, fx=1, fy=1)
    ocr_dict = pytesseract.image_to_data(img, lang='eng', output_type=Output.DICT, config='--psm 4')
    ocr_text = ocr_dict['text']
    while ('' in ocr_text):
        ocr_text.remove('')
    while (' ' in ocr_text):
        ocr_text.remove(' ')
    return ocr_text
def folder2file(path):
    return os.listdir(path)
def files_sort(files, pdf_root_path):
    for i in files:
        page_list = pdf2txt(pdf_root_path+"\\"+str(i))
        print(page_list[0])
        if page_list[0] == []:
            shutil.copy(pdf_root_path + "\\" + str(i), pdf_root_path + "\\" + "OCR\\" + str(i))
        else:
            shutil.copy(pdf_root_path + "\\" + str(i), pdf_root_path + "\\" + "PDF\\" + str(i))

def find_between(key1 ,key2, page):
    owner = re.findall(key1 + ".*" + key2, page)
    if owner != []:
        result = owner[0][len(key1):-len(key2)]
    else:
        result = []
    return result

def main():
    try:

        pdf_path = r"E:\~Ridheesh\Firstplanit\webscrape\epd-norge"
        files = folder2file(pdf_path)
        # file = '1588_Natural-stone-quartzite-schist--even-thickness--with-broken-or-sawn-edges_no.pdf'

        row = 200
        count = len(files)
        while row < len(files):
        # while row < 2:
            try:
                # 0
                print(row, "/",count)
                file = files[row-1]
                file_path = pdf_path+'\\'+file

                with pdfplumber.open(file_path) as pdf:
                    page = pdf.pages[1]
                    left = page.crop((0, 0, 0.45 * page.width, 0.9 * page.height))
                    right = page.crop((0.45 * page.width, 0, page.width, page.height))
                    l_text = left.extract_text()
                    r_text = right.extract_text()
                    page_text = l_text + "\n" + r_text

                page_text_lst = list(page_text.split("\n"))

                # 1
                print(file)
                pdf_file_name = sheet.cell(row, 1)
                pdf_file_name.value = str(file)

                # 2
                keywords = ['Product:', 'Product', 'Product: ', 'Product ','Produkt:', 'Produkt', 'Produkt: ', 'Produkt ']
                for word in keywords:
                    if word in page_text_lst:
                        product_name = page_text_lst[page_text_lst.index(word) + 1]
                        break
                    else:
                        product_name = "-"
                print(product_name, " : " ,file)
                f_prod_name = sheet.cell(row, 2)
                f_prod_name.value = str(product_name)

                # 3
                keywords = ['Program holder:', 'Program holder', 'Program holder ', 'Program holder: ',
                            'Produkt ', 'Produkt: ', 'Produkt', 'Produkt:',
                            'Program operator:', 'Program Operator: ', 'Program Operator ', 'Program Operator',
                            'Program operatør:', 'Progroapme ratør: ', 'Programoperatør', 'Programoperatør: ',  'Programoperatør:',
                            'Programoperatör:',  'Programoperatör', 'Program operatör:',
                            'General Information']
                for word in keywords:
                    if word in page_text_lst:
                        epd_provider = page_text_lst[page_text_lst.index(word) + 1]
                        break
                    else:
                        epd_provider = "-"
                print(epd_provider, " : ", file)
                f_epd_provider = sheet.cell(row, 3)
                f_epd_provider.value = str(epd_provider)

                # 4 5 6
                keywords = ['Eier av deklarasjonen', 'Eier av deklarasjonen:', 'Eier av deklarasjonen: ',  'Eier av deklarasjonen ',
                            'Deklarationens ägare:', 'Deklarationens ägare: ', 'Deklarationens ägare', 'Deklarationens ägare ',
                            'Owner of the declaration:', 'Owner of the declaration: ', 'Owner of the declaration ', 'Owner of the declaration' ,
                            'Ägare av deklarationen: ', 'Ägare av deklarationen ', 'Ägare av deklarationen:',  'Ägare av deklarationen']
                for word in keywords:
                    if word in page_text_lst:
                        owner = page_text_lst[page_text_lst.index(word) + 1]
                        owner_contact = page_text_lst[page_text_lst.index(word) + 2]
                        owner_contact2 = page_text_lst[page_text_lst.index(word) + 3]
                        owner_contact3 = page_text_lst[page_text_lst.index(word) + 4]
                        break
                    else:
                        owner = "-"
                        owner_contact = "-"
                        owner_contact2 = "-"
                        owner_contact3 = "-"
                print(owner, " : ", file)
                print(owner_contact, " : ", file)
                print(owner_contact2, " : ", file)
                print(owner_contact3, " : ", file)
                f_owner = sheet.cell(row, 4)
                f_owner.value = str(owner)
                f_owner_contact = sheet.cell(row, 5)
                f_owner_contact.value = str(owner_contact)
                f_owner_contact2 = sheet.cell(row, 6)
                f_owner_contact2.value = str(owner_contact2)
                f_owner_contact3 = sheet.cell(row, 7)
                f_owner_contact3.value = str(owner_contact3)

                # 7
                keywords = ['Manufacturer:', 'Manufacturer: ', 'Manufacturer', 'Manufacturer ',
                            'Supplier:',  'Supplier: ',  'Supplier',  'Supplier ',
                            'Produsent:', 'Produsent: ', 'Produsent', 'Produsent ',
                            'Producent:','Producent: ','Producent','Producent ',
                            'Tillverkare:', 'Tillverkare: ','Tillverkare','Tillverkare ',
                            'Produsenter:', 'Produsenter: ','Produsenter','Produsenter ',]
                for word in keywords:
                    if word in page_text_lst:
                        manufacturer = page_text_lst[page_text_lst.index(word) + 1]
                        break
                    else:
                        manufacturer = "-"
                print(manufacturer, " : ", file)
                f_manufacturer = sheet.cell(row, 8)
                f_manufacturer.value = str(manufacturer)

                # 8
                keywords = ['Place of production:', 'Place of production: ', 'Place of production','Place of production ',
                            'Produksjonssted:', 'Produksjonssted: ','Produksjonssted','Produksjonssted ',
                            'Produktionsort:', 'Produktionsort: ','Produktionsort ','Produktionsort',
                            'Produksjonssted: ', 'Produksjonssted:', 'Produksjonssted','Produksjonssted ',
                            'Produktionssted:', 'Produktionssted: ', 'Produktionssted','Produktionssted ',
                            'Ort för tillverkning:','Ort för tillverkning: ','Ort för tillverkning','Ort för tillverkning ',]
                for word in keywords:
                    if word in page_text_lst:
                        manu_location = page_text_lst[page_text_lst.index(word) + 1]
                        break
                    else:
                        manu_location = "-"
                print(manu_location, " : ", file)
                f_manu_location = sheet.cell(row, 9)
                f_manu_location.value = str(manu_location)

                # 9 10
                match = re.findall('\d{2}\.\d{2}\.\d{4}', page_text)
                if len(match) >= 2:
                    i_date = match[0]
                    e_date = match[1]
                    print(i_date)
                    f_i_date = sheet.cell(row, 10)
                    f_i_date.value = str(i_date)
                    print(e_date)
                    f_e_date = sheet.cell(row, 11)
                    f_e_date.value = str(e_date)
                else:
                    print("-")
                    f_i_date = sheet.cell(row, 10)
                    f_i_date.value = str("-")
                    print("-")
                    f_e_date = sheet.cell(row, 11)
                    f_e_date.value = str("-")

                # Table Values
                pages = list(pdf2txt(file_path))
                t = 0
                for p in pages:
                    try:
                        p_str = ' '.join(p)
                        if re.findall(".*GWP.*AP.*EP.*", p_str) and re.findall(".*\dE\+|\-\d.*", p_str): #NEED TO OPTIMIZE
                            tables = pdf2tbl(file_path, t + 1)
                            for table in tables:
                                try:
                                    if 'GWP' in table.df.values and ('A1-A3' in table.df.values or 'A1-A3 Nat.' in table.df.values):
                                        data_table = table
                                        table_list = data_table.df.values.tolist()

                                        #Units might be values!
                                        for c, r in enumerate(table_list):
                                            if 'GWP' in r:
                                                gwp_i = c, r.index("GWP")
                                            if "A1-A3" in r:
                                                index = c, r.index("A1-A3")
                                            if "AP" in r:
                                                ap_i = c, r.index("AP")
                                            if "EP" in r:
                                                ep_i = c, r.index("EP")

                                        gwp = table_list[gwp_i[0]][index[1]]
                                        gwp_u = table_list[gwp_i[0]][index[1]-1]
                                        print(gwp, gwp_u)

                                        f_gwp = sheet.cell(row, 12)
                                        f_gwp.value = str(gwp)
                                        f_gwp_u = sheet.cell(row, 13)
                                        f_gwp_u.value = str(gwp_u)

                                        ap = table_list[ap_i[0]][index[1]]
                                        ap_u = table_list[ap_i[0]][index[1] - 1]
                                        print(ap, ap_u)

                                        f_gwp = sheet.cell(row, 14)
                                        f_gwp.value = str(ap)
                                        f_gwp_u = sheet.cell(row, 15)
                                        f_gwp_u.value = str(ap_u)

                                        ep = table_list[ep_i[0]][index[1]]
                                        ep_u = table_list[ep_i[0]][index[1] - 1]
                                        print(ep, ep_u)

                                        f_gwp = sheet.cell(row, 16)
                                        f_gwp.value = str(ep)
                                        f_gwp_u = sheet.cell(row, 17)
                                        f_gwp_u.value = str(ep_u)
                                    elif 'GWP' in table.df.values and 'A1' in table.df.values:
                                        flag = True
                                        data_table = table
                                        table_list = data_table.df.values.tolist()
                                        for c, r in enumerate(table_list):
                                            if 'GWP' in r:
                                                gwp_i = c, r.index("GWP")
                                            if "A1" in r:
                                                index1 = c, r.index("A1")
                                            if "A2" in r:
                                                index2 = c, r.index("A2")
                                            if "A3" in r:
                                                index3 = c, r.index("A3")
                                            if "AP" in r:
                                                ap_i = c, r.index("AP")
                                            if "EP" in r:
                                                ep_i = c, r.index("EP")

                                        gwp1 = table_list[gwp_i[0]][index1[1]]
                                        gwp1_u = table_list[gwp_i[0]][index1[1] - 1]
                                        gwp2 = table_list[gwp_i[0]][index2[1]]
                                        gwp3 = table_list[gwp_i[0]][index3[1]]
                                        print(gwp1, gwp1_u, gwp2, gwp3)
                                        f_gwp1 = sheet.cell(row, 18)
                                        f_gwp1.value = str(gwp1)
                                        f_gwp2 = sheet.cell(row, 19)
                                        f_gwp2.value = str(gwp2)
                                        f_gwp3 = sheet.cell(row, 20)
                                        f_gwp3.value = str(gwp3)
                                        f_gwpu = sheet.cell(row, 21)
                                        f_gwpu.value = str(gwp1_u)

                                        ap1 = table_list[ap_i[0]][index1[1]]
                                        ap1_u = table_list[ap_i[0]][index1[1] - 1]
                                        ap2 = table_list[ap_i[0]][index2[1]]
                                        ap3 = table_list[ap_i[0]][index3[1]]
                                        print(ap1, ap1_u, ap2, ap3)
                                        f_ap1 = sheet.cell(row, 22)
                                        f_ap1.value = str(ap1)
                                        f_ap2 = sheet.cell(row, 23)
                                        f_ap2.value = str(ap2)
                                        f_ap3 = sheet.cell(row, 24)
                                        f_ap3.value = str(ap3)
                                        f_apu = sheet.cell(row, 25)
                                        f_apu.value = str(ap1_u)

                                        ep1 = table_list[ep_i[0]][index1[1]]
                                        ep1_u = table_list[ep_i[0]][index1[1] - 1]
                                        ep2 = table_list[ep_i[0]][index2[1]]
                                        ep3 = table_list[ep_i[0]][index3[1]]
                                        print(ep1, ep1_u, ep2, ep3)
                                        f_ep1 = sheet.cell(row, 26)
                                        f_ep1.value = str(ep1)
                                        f_ep2 = sheet.cell(row, 27)
                                        f_ep2.value = str(ep2)
                                        f_ep3 = sheet.cell(row, 28)
                                        f_ep3.value = str(ep3)
                                        f_epu = sheet.cell(row, 29)
                                        f_epu.value = str(ep1_u)

                                except Exception as e:
                                    wb.save("demo.xlsx")
                                    print("MM-ERROR: " + str(e) + " " + "\n")
                                    continue
                        t += 1
                    except Exception as e:
                        wb.save("demo.xlsx")
                        t += 1
                        print("M-ERROR: " + str(e) + " " + "\n")
                        continue

                wb.save("demo.xlsx")
                print()
                row += 1
            except Exception as e:
                # print(page_text)
                wb.save("demo.xlsx")
                row += 1
                print("ERROR: " + str(e) + " "+ "\n")
                print()
                continue
            # except ZeroDivisionError:
            #     print("Canit divide by 0")

    # except Exception as e:
    #     print("ERROR: " + str(e))
    except ZeroDivisionError:
        print("Canit divide by 0")

if __name__=="__main__":
    main()
#522